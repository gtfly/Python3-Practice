import re
from urllib import parse
import requests
from openpyxl import Workbook
import time
import random

# 返回url内容
def ReadUrl(url):
    USER_AGENTS = [
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; AcooBrowser; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
    "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
    "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
    "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
    "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
    "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
    "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
    "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
    "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/2.0 Safari/536.11",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; LBBROWSER)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E; LBBROWSER)",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.84 Safari/535.11 LBBROWSER",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E)",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; QQBrowser/7.0.3698.400)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; 360SE)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E)",
    "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.89 Safari/537.1",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.89 Safari/537.1",
    "Mozilla/5.0 (iPad; U; CPU OS 4_2_1 like Mac OS X; zh-cn) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8C148 Safari/6533.18.5",
    "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:2.0b13pre) Gecko/20110307 Firefox/4.0b13pre",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:16.0) Gecko/20100101 Firefox/16.0",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11",
    "Mozilla/5.0 (X11; U; Linux x86_64; zh-CN; rv:1.9.2.10) Gecko/20100922 Ubuntu/10.10 (maverick) Firefox/3.6.10"
    ]
 
    headers = {
        'User-Agent': random.choice(USER_AGENTS),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Connection': 'keep-alive',
        'Accept-Encoding': 'gzip, deflate',
    }

    
    proxy_list = ['https://112.85.172.149:9999', 'https://122.193.247.219:9999', 'https://122.193.246.173:9999', 'https://163.204.245.53:9999']
    
    proxies = {'https': random.choice(proxy_list)}
    
    
    content = requests.get(url, headers=headers, proxies=proxies)
    content.encoding = 'gb2312'
    
    return content



# 0级爬取，获取省名称、下级链接，以键值对格式存储在dict1中
def Spider0(url):
    global dict1 
    
    content_obj = ReadUrl(url)
    content = content_obj.text
    
    links = re.findall("href='(.*?html)", content)
    new_links = []
    for i in links:
        i = parse.urljoin(url, i)
        new_links.append(i)
    city = re.findall(".html'>(.*?)<br", content)
    
    dict1 = dict(zip(city, new_links))

    content_obj.close()
    #print(dict1)


# 4级爬取，获取省名称、市名称、县名称、镇名称、居委会名称、划分代码
def Spider4(city, shi_name, xian_name, xian_url, zhen_name, zhen_url):
    global sheet4
    
    zhen_url = parse.urljoin(xian_url, zhen_url)

    time.sleep(1)
    content_obj = ReadUrl(zhen_url)
    content = content_obj.text
    
    res = re.findall(r"<td>(.*?)</td><td>.*?</td><td>(.*?)</td>", content)
    #print(res)
    for i in res:
        data = []
        data.append(city)
        data.append(shi_name)
        data.append(xian_name)
        data.append(zhen_name)
        data.append(i[1])
        data.append(i[0])
        print(data)
        #sheet3.append(data)
        
        
    content_obj.close()


# 3级爬取，获取省名称、市名称、县名称、镇名称、镇划分代码
def Spider3(city, shi_name, shi_url, xian_url, xian_name):
    global sheet3

    xian_url = parse.urljoin(shi_url, xian_url)

    time.sleep(1)
    content_obj = ReadUrl(xian_url)
    content = content_obj.text
        
    res = re.findall(r"href='(.*?)'>(.*?)</a>.*?'>(.*?)</a>", content)
    #print(res)
    for i in res:
        data = []
        data.append(city)
        data.append(shi_name)
        data.append(xian_name)
        data.append(i[2])
        data.append(i[1])
        #print(data)
        #sheet3.append(data)
        Spider4(city, shi_name, xian_name, xian_url, i[2], i[0])
        print(city, shi_name, xian_name, '四级爬取成功！')
        
    content_obj.close()

# 2级爬取，获取省名称、市名称、县名称、县代码，存储在spider2.xlsx文件
def Spider2(city, sheng_url, shi_url, shi_name):
    global sheet2
    
    shi_url = parse.urljoin(sheng_url, shi_url)  # 将相对路径转换为绝对路径

    time.sleep(1)
    content_obj = ReadUrl(shi_url)
    content = content_obj.text
    if '市辖区' in content:   # 市的市辖区没有url，特殊处理
        data = re.findall('<td>(.*?)</td><td>市辖区</td>', content)
        data.insert(0, '市辖区')
        data.insert(0, shi_name)
        data.insert(0, city)
        #print(data)
        sheet2.append(data)
        
    res = re.findall(r"href='(.*?)'>(.*?)</a>.*?'>(.*?)</a>", content)  # [('01/110101.html', '110101000000', '东城区')
    for i in res:
        data = []
        data.append(city)
        data.append(shi_name) 
        data.append(i[2])
        data.append(i[1])
        #print(data)
        #sheet2.append(data)
        Spider3(city, shi_name, shi_url, i[0], i[2])
        print(city, shi_name, '三级爬取成功！')
        
    content_obj.close()

# 1级爬取，获取各市名称及市的代码，存储在spider1.xlsx文件
def Spider1(dict1):
    global sheet1
    
    for city,url in dict1.items():
        
        content_obj = ReadUrl(url)
        content = content_obj.text
        res = re.findall(r"href='(.*?)'>(.*?)</a>.*?'>(.*?)</a>", content)  # [('51/5101.html', '510100000000', '成都市')]
        
        for i in res:
            #print(i)
            data = []
            data.append(city)
            data.append(i[2])
            data.append(i[1])
            #sheet1.append(data)
            Spider2(city, url, i[0], i[2])  # 将省名称、省url、市url、市名称传递给二级爬虫
        content_obj.close()
        print(city, '二级爬取成功！')
        
    print('二级爬取成功！')
    
 
if __name__ == '__main__':

    print('开始时间：', time.ctime())
    startUrl = 'http://www.stats.gov.cn/tjsj/tjbz/tjyqhdmhcxhfdm/2018/index.html'
    
    
    wb1 = Workbook()
    sheet1 = wb1.active
    sheet1.append(['省', '市', '统计用区划代码'])
    
    
    wb2 = Workbook()
    sheet2 = wb2.active
    sheet2.append(['省', '市', '县', '统计用区划代码'])
    
    
    wb3 = Workbook()
    sheet3 = wb3.active
    sheet2.append(['省', '市', '县', '镇', '统计用区划代码'])


    wb4 = Workbook()
    sheet4 = wb4.active
    sheet2.append(['省', '市', '县', '镇', '居委会', '统计用区划代码'])

    
    # 爬取
    dict1 = {}
    Spider0(startUrl)
    #print(dict1)
    
    Spider1(dict1)
    #Spider1(dict1)
    
    
    # 生成文件
    #wb1.save('C:/Users/腾飞/Desktop/spider1.xlsx')
    
    #wb2.save('C:/Users/腾飞/Desktop/spider2.xlsx')
    
    
    #wb3.save('C:/Users/腾飞/Desktop/spider3.xlsx')
    #wb4.save('C:/Users/腾飞/Desktop/spider4.xlsx')
    
    print('结束时间：', time.ctime())
    

